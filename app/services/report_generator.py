# -*- coding: utf-8 -*-
"""
KZ-InsurePro - Генератор регуляторных отчётов
Полноценный экспорт в Excel с форматированием, PDF, отправка в АРРФР

Поддерживаемые формы:
- Постановление №85: 1-СК, 2-СК, 3-СК, 4-СК, 5-СК
- Постановление №86: П-86/1, П-86/2, П-86/3
- Постановление №304: Р-304/1, Р-304/2, Р-304/3, Р-304/4
"""

import io
from datetime import date, datetime
from decimal import Decimal
from typing import Dict, List, Any, Optional
from dataclasses import dataclass
from enum import Enum
import json

# Excel export
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, NamedStyle
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# PDF export (optional)
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False


class ReportFormat(Enum):
    """Форматы экспорта"""
    EXCEL = "xlsx"
    PDF = "pdf"
    JSON = "json"
    XML = "xml"


@dataclass
class ReportMetadata:
    """Метаданные отчёта"""
    report_code: str
    report_name: str
    report_name_kz: str
    regulation: str
    period: str
    period_start: date
    period_end: date
    company_name: str
    company_bin: str
    generated_at: datetime
    generated_by: str


@dataclass
class GeneratedReport:
    """Сгенерированный отчёт"""
    metadata: ReportMetadata
    data: Dict[str, Any]
    content: bytes  # Binary content (Excel/PDF)
    format: ReportFormat
    filename: str


class ReportGeneratorService:
    """
    Сервис генерации регуляторных отчётов АРРФР

    Генерирует полноценные отчёты с форматированием по формам АРРФР.
    """

    def __init__(self):
        self.company_info = {
            'name': 'ТОО "Страховая компания Казахстан"',
            'name_kz': '"Қазақстан сақтандыру компаниясы" ЖШС',
            'bin': '123456789012',
            'license': 'СК-001',
            'license_date': '15.03.2020',
            'address': 'г. Алматы, ул. Достык, 1',
            'phone': '+7 (727) 123-45-67',
            'email': 'info@insurance.kz',
            'director': 'Иванов Иван Иванович',
            'accountant': 'Петрова Анна Сергеевна',
            'actuary': 'Сидоров Пётр Михайлович'
        }

        # Демо-данные для отчётов
        self.demo_data = self._load_demo_data()

    def _load_demo_data(self) -> Dict:
        """Загрузка демо-данных для отчётов"""
        return {
            'premiums': {
                'ogpo_vts': Decimal('1234567000'),
                'kasko': Decimal('456789000'),
                'property_corp': Decimal('345678000'),
                'property_ind': Decimal('123456000'),
                'life': Decimal('234567000'),
                'health': Decimal('178901000'),
                'gpo': Decimal('97330000'),
                'cargo': Decimal('67890000'),
                'liability': Decimal('45678000'),
                'total': Decimal('2784856000')
            },
            'claims': {
                'ogpo_vts': Decimal('567890000'),
                'kasko': Decimal('234567000'),
                'property_corp': Decimal('123456000'),
                'property_ind': Decimal('56789000'),
                'life': Decimal('89012000'),
                'health': Decimal('134567000'),
                'gpo': Decimal('85075000'),
                'cargo': Decimal('23456000'),
                'liability': Decimal('12345000'),
                'total': Decimal('1327157000')
            },
            'contracts': {
                'ogpo_vts': 23456,
                'kasko': 8765,
                'property_corp': 1234,
                'property_ind': 4567,
                'life': 5678,
                'health': 4321,
                'gpo': 2224,
                'cargo': 567,
                'liability': 345,
                'total': 51157
            },
            'reserves': {
                'rnpp': Decimal('1523015000'),
                'rzu': Decimal('1116418000'),
                'ibnr': Decimal('536417000'),
                'rpzu': Decimal('101493000'),
                'stabilization': Decimal('250000000'),
                'catastrophe': Decimal('150000000'),
                'total': Decimal('3677343000')
            },
            'capital': {
                'authorized': Decimal('5000000000'),
                'additional': Decimal('2500000000'),
                'reserve': Decimal('1250000000'),
                'retained': Decimal('3750000000'),
                'revaluation': Decimal('500000000'),
                'subordinated': Decimal('2500000000'),
                'total': Decimal('15500000000')
            },
            'assets': {
                'government_securities': Decimal('8500000000'),
                'bank_deposits': Decimal('6250000000'),
                'corporate_bonds': Decimal('4500000000'),
                'stocks': Decimal('2750000000'),
                'real_estate': Decimal('2000000000'),
                'receivables': Decimal('1500000000'),
                'cash': Decimal('750000000'),
                'other': Decimal('500000000'),
                'total': Decimal('26750000000')
            },
            'solvency': {
                'fmp': Decimal('13300000000'),
                'nmp': Decimal('6957386000'),
                'mgf': Decimal('2319129000'),
                'ratio': Decimal('1.912'),
                'surplus': Decimal('6342614000')
            },
            'reinsurance': {
                'ceded_premiums': Decimal('695000000'),
                'ceded_claims': Decimal('331789000'),
                'reinsurance_commission': Decimal('104250000'),
                'retention_ratio': Decimal('0.75')
            }
        }

    def generate_report(
        self,
        report_code: str,
        period: str,
        format: ReportFormat = ReportFormat.EXCEL
    ) -> GeneratedReport:
        """
        Генерация отчёта

        Args:
            report_code: Код формы (1-SK, P86-1, R304-1 и т.д.)
            period: Период (2026-01, 2025-Q4, 2025 и т.д.)
            format: Формат экспорта

        Returns:
            GeneratedReport
        """
        # Parse period
        period_start, period_end = self._parse_period(period)

        # Create metadata
        metadata = ReportMetadata(
            report_code=report_code,
            report_name=self._get_report_name(report_code),
            report_name_kz=self._get_report_name_kz(report_code),
            regulation=self._get_regulation(report_code),
            period=period,
            period_start=period_start,
            period_end=period_end,
            company_name=self.company_info['name'],
            company_bin=self.company_info['bin'],
            generated_at=datetime.now(),
            generated_by='system'
        )

        # Generate report data
        data = self._generate_report_data(report_code, period_start, period_end)

        # Export to format
        if format == ReportFormat.EXCEL:
            content, filename = self._export_to_excel(report_code, metadata, data)
        elif format == ReportFormat.PDF:
            content, filename = self._export_to_pdf(report_code, metadata, data)
        elif format == ReportFormat.JSON:
            content, filename = self._export_to_json(report_code, metadata, data)
        else:
            content, filename = self._export_to_excel(report_code, metadata, data)

        return GeneratedReport(
            metadata=metadata,
            data=data,
            content=content,
            format=format,
            filename=filename
        )

    def _parse_period(self, period: str) -> tuple:
        """Парсинг периода"""
        if '-Q' in period:
            # Квартал: 2025-Q4
            year, quarter = period.split('-Q')
            quarter = int(quarter)
            if quarter == 1:
                return date(int(year), 1, 1), date(int(year), 3, 31)
            elif quarter == 2:
                return date(int(year), 4, 1), date(int(year), 6, 30)
            elif quarter == 3:
                return date(int(year), 7, 1), date(int(year), 9, 30)
            else:
                return date(int(year), 10, 1), date(int(year), 12, 31)
        elif len(period) == 7:
            # Месяц: 2026-01
            year, month = period.split('-')
            year, month = int(year), int(month)
            if month == 12:
                return date(year, month, 1), date(year, 12, 31)
            else:
                next_month = date(year, month + 1, 1)
                last_day = date(next_month.year, next_month.month, 1)
                from datetime import timedelta
                return date(year, month, 1), last_day - timedelta(days=1)
        else:
            # Год: 2025
            year = int(period)
            return date(year, 1, 1), date(year, 12, 31)

    def _get_report_name(self, code: str) -> str:
        """Получение названия отчёта"""
        names = {
            '1-SK': 'Отчёт о страховой деятельности',
            '2-SK': 'Отчёт о страховых премиях и выплатах по классам страхования',
            '3-SK': 'Отчёт о структуре активов страховой организации',
            '4-SK': 'Отчёт о перестраховочных операциях',
            '5-SK': 'Годовой бухгалтерский баланс',
            'P86-1': 'Расчёт маржи платёжеспособности страховой организации',
            'P86-2': 'Расчёт минимального гарантийного фонда',
            'P86-3': 'Расчёт пруденциальных нормативов достаточности капитала',
            'R304-1': 'Расчёт резерва незаработанной премии (РНПП)',
            'R304-2': 'Расчёт резерва заявленных, но неурегулированных убытков (РЗУ)',
            'R304-3': 'Расчёт резерва произошедших, но незаявленных убытков (IBNR)',
            'R304-4': 'Расчёт стабилизационного резерва'
        }
        return names.get(code, f'Форма {code}')

    def _get_report_name_kz(self, code: str) -> str:
        """Получение названия на казахском"""
        names = {
            '1-SK': 'Сақтандыру қызметі туралы есеп',
            '2-SK': 'Сақтандыру сыйлықақылары мен төлемдері туралы есеп',
            '3-SK': 'Активтер құрылымы туралы есеп',
            '4-SK': 'Қайта сақтандыру операциялары туралы есеп',
            '5-SK': 'Жылдық бухгалтерлік баланс',
            'P86-1': 'Төлем қабілеттілігі маржасын есептеу',
            'P86-2': 'Ең төменгі кепілдік қорын есептеу',
            'P86-3': 'Капитал жеткіліктілігі нормативтерін есептеу',
            'R304-1': 'Жұмыс істемеген сыйлықақы резервін есептеу (РНПП)',
            'R304-2': 'Мәлімделген шығындар резервін есептеу (РЗУ)',
            'R304-3': 'IBNR резервін тізбекті баспалдақ әдісімен есептеу',
            'R304-4': 'Тұрақтандыру резервін есептеу'
        }
        return names.get(code, f'{code} нысаны')

    def _get_regulation(self, code: str) -> str:
        """Определение постановления"""
        if code.startswith('P86'):
            return 'Постановление Правления АРРФР №86 от 27.04.2020'
        elif code.startswith('R304'):
            return 'Постановление Правления АРРФР №304 от 25.12.2023'
        else:
            return 'Постановление Правления АРРФР №85 от 27.04.2020'

    def _generate_report_data(self, report_code: str, period_start: date, period_end: date) -> Dict[str, Any]:
        """Генерация данных отчёта"""
        data = {
            'company': self.company_info,
            'period_start': period_start.isoformat(),
            'period_end': period_end.isoformat()
        }

        if report_code == '1-SK':
            data['sections'] = self._generate_1sk_data()
        elif report_code == '2-SK':
            data['sections'] = self._generate_2sk_data()
        elif report_code == '3-SK':
            data['sections'] = self._generate_3sk_data()
        elif report_code == '4-SK':
            data['sections'] = self._generate_4sk_data()
        elif report_code == 'P86-1':
            data['sections'] = self._generate_p861_data()
        elif report_code == 'P86-2':
            data['sections'] = self._generate_p862_data()
        elif report_code == 'P86-3':
            data['sections'] = self._generate_p863_data()
        elif report_code == 'R304-1':
            data['sections'] = self._generate_r3041_data()
        elif report_code == 'R304-2':
            data['sections'] = self._generate_r3042_data()
        elif report_code == 'R304-3':
            data['sections'] = self._generate_r3043_data()
        elif report_code == 'R304-4':
            data['sections'] = self._generate_r3044_data()
        else:
            data['sections'] = []

        return data

    def _generate_1sk_data(self) -> List[Dict]:
        """Данные для формы 1-СК"""
        d = self.demo_data
        return [
            {'row': '1', 'name': 'Страховые премии - всего', 'value': d['premiums']['total'], 'prev_value': d['premiums']['total'] * Decimal('0.92')},
            {'row': '1.1', 'name': '  в том числе по обязательному страхованию', 'value': d['premiums']['ogpo_vts'] + d['premiums']['gpo'], 'prev_value': (d['premiums']['ogpo_vts'] + d['premiums']['gpo']) * Decimal('0.95')},
            {'row': '1.1.1', 'name': '    ОГПО ВТС', 'value': d['premiums']['ogpo_vts'], 'prev_value': d['premiums']['ogpo_vts'] * Decimal('0.93')},
            {'row': '1.1.2', 'name': '    ГПО работодателя', 'value': d['premiums']['gpo'], 'prev_value': d['premiums']['gpo'] * Decimal('0.97')},
            {'row': '1.2', 'name': '  добровольное личное страхование', 'value': d['premiums']['life'] + d['premiums']['health'], 'prev_value': (d['premiums']['life'] + d['premiums']['health']) * Decimal('0.88')},
            {'row': '1.2.1', 'name': '    страхование жизни', 'value': d['premiums']['life'], 'prev_value': d['premiums']['life'] * Decimal('0.85')},
            {'row': '1.2.2', 'name': '    медицинское страхование', 'value': d['premiums']['health'], 'prev_value': d['premiums']['health'] * Decimal('0.91')},
            {'row': '1.3', 'name': '  добровольное имущественное страхование', 'value': d['premiums']['kasko'] + d['premiums']['property_corp'] + d['premiums']['property_ind'], 'prev_value': (d['premiums']['kasko'] + d['premiums']['property_corp'] + d['premiums']['property_ind']) * Decimal('0.89')},
            {'row': '1.3.1', 'name': '    КАСКО', 'value': d['premiums']['kasko'], 'prev_value': d['premiums']['kasko'] * Decimal('0.87')},
            {'row': '1.3.2', 'name': '    страхование имущества юр.лиц', 'value': d['premiums']['property_corp'], 'prev_value': d['premiums']['property_corp'] * Decimal('0.91')},
            {'row': '1.3.3', 'name': '    страхование имущества физ.лиц', 'value': d['premiums']['property_ind'], 'prev_value': d['premiums']['property_ind'] * Decimal('0.94')},
            {'row': '2', 'name': 'Страховые выплаты - всего', 'value': d['claims']['total'], 'prev_value': d['claims']['total'] * Decimal('0.88')},
            {'row': '2.1', 'name': '  в том числе по обязательному страхованию', 'value': d['claims']['ogpo_vts'] + d['claims']['gpo'], 'prev_value': (d['claims']['ogpo_vts'] + d['claims']['gpo']) * Decimal('0.90')},
            {'row': '2.2', 'name': '  добровольное личное страхование', 'value': d['claims']['life'] + d['claims']['health'], 'prev_value': (d['claims']['life'] + d['claims']['health']) * Decimal('0.85')},
            {'row': '2.3', 'name': '  добровольное имущественное страхование', 'value': d['claims']['kasko'] + d['claims']['property_corp'] + d['claims']['property_ind'], 'prev_value': (d['claims']['kasko'] + d['claims']['property_corp'] + d['claims']['property_ind']) * Decimal('0.87')},
            {'row': '3', 'name': 'Количество заключённых договоров', 'value': d['contracts']['total'], 'prev_value': int(d['contracts']['total'] * 0.94)},
            {'row': '4', 'name': 'Количество урегулированных убытков', 'value': 18567, 'prev_value': 16234},
            {'row': '5', 'name': 'Среднее время урегулирования убытка (дней)', 'value': 12, 'prev_value': 14},
            {'row': '6', 'name': 'Коэффициент убыточности (%)', 'value': Decimal('47.65'), 'prev_value': Decimal('45.23')},
            {'row': '7', 'name': 'Комбинированный коэффициент (%)', 'value': Decimal('89.45'), 'prev_value': Decimal('91.23')},
        ]

    def _generate_2sk_data(self) -> List[Dict]:
        """Данные для формы 2-СК"""
        d = self.demo_data
        classes = [
            ('Обязательное страхование гражданско-правовой ответственности владельцев транспортных средств (ОГПО ВТС)', 'ogpo_vts'),
            ('Добровольное страхование автотранспорта (КАСКО)', 'kasko'),
            ('Страхование имущества юридических лиц', 'property_corp'),
            ('Страхование имущества физических лиц', 'property_ind'),
            ('Страхование жизни', 'life'),
            ('Добровольное медицинское страхование', 'health'),
            ('ГПО работодателя за причинение вреда жизни и здоровью работника', 'gpo'),
            ('Страхование грузов', 'cargo'),
            ('Страхование гражданско-правовой ответственности', 'liability'),
        ]

        rows = []
        for i, (name, key) in enumerate(classes, 1):
            premium = d['premiums'].get(key, Decimal('0'))
            claims = d['claims'].get(key, Decimal('0'))
            contracts = d['contracts'].get(key, 0)
            loss_ratio = (claims / premium * 100) if premium > 0 else Decimal('0')

            rows.append({
                'row': str(i),
                'class_name': name,
                'premiums_gross': premium,
                'premiums_net': premium * Decimal('0.75'),  # За вычетом перестрахования
                'claims_gross': claims,
                'claims_net': claims * Decimal('0.75'),
                'contracts': contracts,
                'loss_ratio': loss_ratio
            })

        # Итого
        rows.append({
            'row': '',
            'class_name': 'ИТОГО',
            'premiums_gross': d['premiums']['total'],
            'premiums_net': d['premiums']['total'] * Decimal('0.75'),
            'claims_gross': d['claims']['total'],
            'claims_net': d['claims']['total'] * Decimal('0.75'),
            'contracts': d['contracts']['total'],
            'loss_ratio': (d['claims']['total'] / d['premiums']['total'] * 100) if d['premiums']['total'] > 0 else Decimal('0')
        })

        return rows

    def _generate_3sk_data(self) -> List[Dict]:
        """Данные для формы 3-СК - структура активов"""
        d = self.demo_data
        total = d['assets']['total']

        return [
            {'row': '1', 'asset_type': 'Государственные ценные бумаги Республики Казахстан',
             'amount': d['assets']['government_securities'],
             'share': d['assets']['government_securities'] / total * 100,
             'limit': Decimal('100'), 'status': 'OK', 'note': 'Без ограничений'},
            {'row': '2', 'asset_type': 'Депозиты в банках второго уровня с рейтингом не ниже BB-',
             'amount': d['assets']['bank_deposits'],
             'share': d['assets']['bank_deposits'] / total * 100,
             'limit': Decimal('40'), 'status': 'OK', 'note': 'Не более 40%'},
            {'row': '3', 'asset_type': 'Корпоративные облигации с рейтингом не ниже BB',
             'amount': d['assets']['corporate_bonds'],
             'share': d['assets']['corporate_bonds'] / total * 100,
             'limit': Decimal('30'), 'status': 'OK', 'note': 'Не более 30%'},
            {'row': '4', 'asset_type': 'Акции, включённые в официальный список KASE (категории A, B)',
             'amount': d['assets']['stocks'],
             'share': d['assets']['stocks'] / total * 100,
             'limit': Decimal('15'), 'status': 'OK', 'note': 'Не более 15%'},
            {'row': '5', 'asset_type': 'Недвижимость (здания, сооружения)',
             'amount': d['assets']['real_estate'],
             'share': d['assets']['real_estate'] / total * 100,
             'limit': Decimal('15'), 'status': 'OK', 'note': 'Не более 15%'},
            {'row': '6', 'asset_type': 'Дебиторская задолженность по страховой деятельности',
             'amount': d['assets']['receivables'],
             'share': d['assets']['receivables'] / total * 100,
             'limit': Decimal('10'), 'status': 'OK', 'note': 'Не более 10%'},
            {'row': '7', 'asset_type': 'Денежные средства',
             'amount': d['assets']['cash'],
             'share': d['assets']['cash'] / total * 100,
             'limit': Decimal('100'), 'status': 'OK', 'note': 'Без ограничений'},
            {'row': '8', 'asset_type': 'Прочие активы',
             'amount': d['assets']['other'],
             'share': d['assets']['other'] / total * 100,
             'limit': Decimal('5'), 'status': 'OK', 'note': 'Не более 5%'},
            {'row': '', 'asset_type': 'ИТОГО АКТИВЫ',
             'amount': total,
             'share': Decimal('100'),
             'limit': None, 'status': '', 'note': ''}
        ]

    def _generate_4sk_data(self) -> Dict:
        """Данные для формы 4-СК - перестрахование"""
        d = self.demo_data
        return {
            'summary': [
                {'row': '1', 'name': 'Премии, переданные в перестрахование', 'value': d['reinsurance']['ceded_premiums']},
                {'row': '2', 'name': 'Доля перестраховщиков в выплатах', 'value': d['reinsurance']['ceded_claims']},
                {'row': '3', 'name': 'Перестраховочная комиссия полученная', 'value': d['reinsurance']['reinsurance_commission']},
                {'row': '4', 'name': 'Коэффициент собственного удержания', 'value': d['reinsurance']['retention_ratio']},
            ],
            'by_class': [
                {'class': 'ОГПО ВТС', 'ceded': Decimal('185000000'), 'recovered': Decimal('85000000')},
                {'class': 'КАСКО', 'ceded': Decimal('137000000'), 'recovered': Decimal('70000000')},
                {'class': 'Имущество юр.лиц', 'ceded': Decimal('173000000'), 'recovered': Decimal('62000000')},
                {'class': 'Прочие классы', 'ceded': Decimal('200000000'), 'recovered': Decimal('114789000')},
            ],
            'reinsurers': [
                {'name': 'Munich Re', 'country': 'Германия', 'rating': 'AA-', 'share': Decimal('35')},
                {'name': 'Swiss Re', 'country': 'Швейцария', 'rating': 'AA-', 'share': Decimal('30')},
                {'name': 'Hannover Re', 'country': 'Германия', 'rating': 'AA-', 'share': Decimal('20')},
                {'name': 'SCOR', 'country': 'Франция', 'rating': 'AA-', 'share': Decimal('15')},
            ]
        }

    def _generate_p861_data(self) -> Dict:
        """Данные для формы П-86/1 - маржа платёжеспособности"""
        d = self.demo_data
        return {
            'fmp_section': [
                {'row': '1', 'name': 'Уставный капитал (оплаченный)', 'value': d['capital']['authorized']},
                {'row': '2', 'name': 'Дополнительный оплаченный капитал', 'value': d['capital']['additional']},
                {'row': '3', 'name': 'Резервный капитал', 'value': d['capital']['reserve']},
                {'row': '4', 'name': 'Нераспределённая прибыль (непокрытый убыток)', 'value': d['capital']['retained']},
                {'row': '5', 'name': 'Резерв переоценки', 'value': d['capital']['revaluation']},
                {'row': '6', 'name': 'Субординированный долг (не более 50% от п.1-5)', 'value': d['capital']['subordinated']},
                {'row': '7', 'name': 'ИТОГО собственный капитал (п.1+п.2+п.3+п.4+п.5+п.6)', 'value': d['capital']['total']},
                {'row': '8', 'name': '(-) Нематериальные активы', 'value': Decimal('-500000000')},
                {'row': '9', 'name': '(-) Отложенные аквизиционные расходы', 'value': Decimal('-1200000000')},
                {'row': '10', 'name': '(-) Дебиторская задолженность свыше 90 дней', 'value': Decimal('-500000000')},
                {'row': '11', 'name': 'ФМП (Фактическая маржа платёжеспособности)', 'value': d['solvency']['fmp'], 'is_total': True},
            ],
            'nmp_section': [
                {'row': '1', 'name': 'Страховые премии за последние 12 месяцев (брутто)', 'value': Decimal('25478320000')},
                {'row': '2', 'name': 'НМП по премиям = п.1 × 18%', 'value': Decimal('4586098000')},
                {'row': '3', 'name': 'Страховые выплаты за последние 36 месяцев (брутто)', 'value': Decimal('35678900000')},
                {'row': '4', 'name': 'НМП по выплатам = п.3 / 3 × 26%', 'value': Decimal('3092171000')},
                {'row': '5', 'name': 'НМП (максимум из п.2 и п.4)', 'value': Decimal('4586098000')},
                {'row': '6', 'name': 'Коэффициент участия перестраховщиков (не менее 0.5)', 'value': Decimal('0.75')},
                {'row': '7', 'name': 'НМП × коэффициент перестрахования (не менее 50%)', 'value': d['solvency']['nmp'], 'is_total': True},
            ],
            'mgf_section': [
                {'row': '1', 'name': 'Минимальный уставный капитал', 'value': Decimal('2000000000')},
                {'row': '2', 'name': '1/3 от НМП', 'value': d['solvency']['nmp'] / 3},
                {'row': '3', 'name': 'МГФ (максимум из п.1 и п.2)', 'value': d['solvency']['mgf'], 'is_total': True},
            ],
            'result': {
                'fmp': d['solvency']['fmp'],
                'nmp': d['solvency']['nmp'],
                'ratio': d['solvency']['ratio'],
                'surplus': d['solvency']['surplus'],
                'mgf': d['solvency']['mgf'],
                'status': 'СООТВЕТСТВУЕТ' if d['solvency']['ratio'] >= 1 else 'НЕ СООТВЕТСТВУЕТ',
                'status_mgf': 'СООТВЕТСТВУЕТ' if d['solvency']['fmp'] >= d['solvency']['mgf'] else 'НЕ СООТВЕТСТВУЕТ'
            }
        }

    def _generate_p862_data(self) -> Dict:
        """Данные для формы П-86/2 - минимальный гарантийный фонд"""
        d = self.demo_data
        return {
            'calculation': [
                {'row': '1', 'name': 'Минимальный уставный капитал для общего страхования', 'value': Decimal('2000000000')},
                {'row': '2', 'name': 'Нормативная маржа платёжеспособности (НМП)', 'value': d['solvency']['nmp']},
                {'row': '3', 'name': '1/3 от НМП', 'value': d['solvency']['nmp'] / 3},
                {'row': '4', 'name': 'МГФ = max(п.1, п.3)', 'value': d['solvency']['mgf'], 'is_total': True},
                {'row': '5', 'name': 'Фактический собственный капитал', 'value': d['solvency']['fmp']},
                {'row': '6', 'name': 'Превышение/недостаток (п.5 - п.4)', 'value': d['solvency']['fmp'] - d['solvency']['mgf'], 'is_total': True},
            ],
            'status': 'СООТВЕТСТВУЕТ' if d['solvency']['fmp'] >= d['solvency']['mgf'] else 'НЕ СООТВЕТСТВУЕТ'
        }

    def _generate_p863_data(self) -> Dict:
        """Данные для формы П-86/3 - пруденциальные нормативы"""
        d = self.demo_data
        return {
            'norms': [
                {'code': 'k1', 'name': 'Достаточность собственного капитала (ФМП/НМП)',
                 'norm': '≥ 1.0', 'actual': d['solvency']['ratio'], 'status': 'OK'},
                {'code': 'k2', 'name': 'Достаточность высоколиквидных активов',
                 'norm': '≥ 0.8', 'actual': Decimal('0.92'), 'status': 'OK'},
                {'code': 'k3', 'name': 'Максимальный размер риска на одного страхователя',
                 'norm': '≤ 10%', 'actual': Decimal('7.5'), 'status': 'OK'},
                {'code': 'k4', 'name': 'Максимальный размер риска на связанных лиц',
                 'norm': '≤ 25%', 'actual': Decimal('18.3'), 'status': 'OK'},
                {'code': 'k5', 'name': 'Коэффициент ликвидности',
                 'norm': '≥ 1.0', 'actual': Decimal('1.45'), 'status': 'OK'},
                {'code': 'k6', 'name': 'Соотношение обязательств к собственному капиталу',
                 'norm': '≤ 5.0', 'actual': Decimal('3.2'), 'status': 'OK'},
            ]
        }

    def _generate_r3041_data(self) -> List[Dict]:
        """Данные для формы Р-304/1 - РНПП"""
        return [
            {'row': '1', 'class': 'ОГПО ВТС', 'premiums': Decimal('1234567000'),
             'inception_date': '01.01.2026', 'expiry_date': '31.12.2026',
             'days_total': 365, 'days_remaining': 213,
             'unearned_ratio': Decimal('58.36'), 'rnpp': Decimal('720355000')},
            {'row': '2', 'class': 'КАСКО', 'premiums': Decimal('456789000'),
             'inception_date': '15.10.2025', 'expiry_date': '14.10.2026',
             'days_total': 365, 'days_remaining': 228,
             'unearned_ratio': Decimal('62.47'), 'rnpp': Decimal('285355000')},
            {'row': '3', 'class': 'Имущество юридических лиц', 'premiums': Decimal('345678000'),
             'inception_date': '01.04.2025', 'expiry_date': '31.03.2026',
             'days_total': 365, 'days_remaining': 59,
             'unearned_ratio': Decimal('16.16'), 'rnpp': Decimal('55861000')},
            {'row': '4', 'class': 'Имущество физических лиц', 'premiums': Decimal('123456000'),
             'inception_date': '01.07.2025', 'expiry_date': '30.06.2026',
             'days_total': 365, 'days_remaining': 151,
             'unearned_ratio': Decimal('41.37'), 'rnpp': Decimal('51082000')},
            {'row': '5', 'class': 'Страхование жизни', 'premiums': Decimal('234567000'),
             'inception_date': '01.01.2025', 'expiry_date': '31.12.2029',
             'days_total': 1826, 'days_remaining': 1430,
             'unearned_ratio': Decimal('78.31'), 'rnpp': Decimal('183681000')},
            {'row': '6', 'class': 'Медицинское страхование', 'premiums': Decimal('178901000'),
             'inception_date': '01.09.2025', 'expiry_date': '31.08.2026',
             'days_total': 365, 'days_remaining': 213,
             'unearned_ratio': Decimal('58.36'), 'rnpp': Decimal('104406000')},
            {'row': '7', 'class': 'ГПО работодателя', 'premiums': Decimal('97330000'),
             'inception_date': '01.01.2026', 'expiry_date': '31.12.2026',
             'days_total': 365, 'days_remaining': 335,
             'unearned_ratio': Decimal('91.78'), 'rnpp': Decimal('89333000')},
            {'row': '8', 'class': 'Страхование грузов', 'premiums': Decimal('67890000'),
             'inception_date': 'Различные', 'expiry_date': 'Различные',
             'days_total': '-', 'days_remaining': '-',
             'unearned_ratio': Decimal('45.00'), 'rnpp': Decimal('30551000')},
            {'row': '', 'class': 'ИТОГО РНПП', 'premiums': Decimal('2739178000'),
             'inception_date': '', 'expiry_date': '',
             'days_total': '', 'days_remaining': '',
             'unearned_ratio': Decimal('55.59'), 'rnpp': Decimal('1520624000'), 'is_total': True},
        ]

    def _generate_r3042_data(self) -> List[Dict]:
        """Данные для формы Р-304/2 - РЗУ"""
        return [
            {'row': '1', 'class': 'ОГПО ВТС', 'claims_count': 1234,
             'estimated_gross': Decimal('567890000'), 'estimated_net': Decimal('425918000'),
             'rpzu': Decimal('56789000'), 'total': Decimal('624679000')},
            {'row': '2', 'class': 'КАСКО', 'claims_count': 456,
             'estimated_gross': Decimal('234567000'), 'estimated_net': Decimal('175925000'),
             'rpzu': Decimal('23457000'), 'total': Decimal('258024000')},
            {'row': '3', 'class': 'Имущество юридических лиц', 'claims_count': 89,
             'estimated_gross': Decimal('123456000'), 'estimated_net': Decimal('92592000'),
             'rpzu': Decimal('12346000'), 'total': Decimal('135802000')},
            {'row': '4', 'class': 'Имущество физических лиц', 'claims_count': 156,
             'estimated_gross': Decimal('56789000'), 'estimated_net': Decimal('42592000'),
             'rpzu': Decimal('5679000'), 'total': Decimal('62468000')},
            {'row': '5', 'class': 'Страхование жизни', 'claims_count': 78,
             'estimated_gross': Decimal('89012000'), 'estimated_net': Decimal('66759000'),
             'rpzu': Decimal('8901000'), 'total': Decimal('97913000')},
            {'row': '6', 'class': 'Медицинское страхование', 'claims_count': 234,
             'estimated_gross': Decimal('134567000'), 'estimated_net': Decimal('100925000'),
             'rpzu': Decimal('13457000'), 'total': Decimal('148024000')},
            {'row': '7', 'class': 'ГПО работодателя', 'claims_count': 45,
             'estimated_gross': Decimal('85075000'), 'estimated_net': Decimal('63806000'),
             'rpzu': Decimal('8508000'), 'total': Decimal('93583000')},
            {'row': '', 'class': 'ИТОГО РЗУ', 'claims_count': 2292,
             'estimated_gross': Decimal('1291356000'), 'estimated_net': Decimal('968517000'),
             'rpzu': Decimal('129137000'), 'total': Decimal('1420493000'), 'is_total': True},
        ]

    def _generate_r3043_data(self) -> Dict:
        """Данные для формы Р-304/3 - IBNR методом Chain-Ladder"""
        return {
            'triangle': [
                {'year': 2022, 'dev0': 234567, 'dev1': 345678, 'dev2': 378901, 'dev3': 389012, 'dev4': 391234, 'ultimate': 391234, 'ibnr': 0},
                {'year': 2023, 'dev0': 256789, 'dev1': 378901, 'dev2': 415234, 'dev3': 427456, 'dev4': None, 'ultimate': 430012, 'ibnr': 2556},
                {'year': 2024, 'dev0': 289012, 'dev1': 426789, 'dev2': 467890, 'dev3': None, 'dev4': None, 'ultimate': 496234, 'ibnr': 28344},
                {'year': 2025, 'dev0': 312345, 'dev1': 461234, 'dev2': None, 'dev3': None, 'dev4': None, 'ultimate': 551678, 'ibnr': 90444},
                {'year': 2026, 'dev0': 156789, 'dev1': None, 'dev2': None, 'dev3': None, 'dev4': None, 'ultimate': 276543, 'ibnr': 119754},
            ],
            'ldf': {
                '0_1': Decimal('1.476'),
                '1_2': Decimal('1.097'),
                '2_3': Decimal('1.030'),
                '3_4': Decimal('1.006'),
                '4_ult': Decimal('1.000')
            },
            'cdf': {
                '0_1': Decimal('1.762'),
                '1_2': Decimal('1.194'),
                '2_3': Decimal('1.089'),
                '3_4': Decimal('1.006'),
                '4_ult': Decimal('1.000')
            },
            'total_ibnr': {
                'ogpo_vts': Decimal('281418000'),
                'kasko': Decimal('156234000'),
                'property': Decimal('67890000'),
                'other': Decimal('30875000'),
                'total': Decimal('536417000')
            }
        }

    def _generate_r3044_data(self) -> Dict:
        """Данные для формы Р-304/4 - стабилизационный резерв"""
        return {
            'by_class': [
                {'class': 'ОГПО ВТС', 'loss_ratio': Decimal('46.0'), 'threshold': Decimal('60'),
                 'deviation': Decimal('0'), 'reserve': Decimal('0'), 'note': 'Убыточность в норме'},
                {'class': 'Страхование от катастрофических рисков', 'loss_ratio': Decimal('15.0'), 'threshold': Decimal('40'),
                 'deviation': Decimal('0'), 'reserve': Decimal('150000000'), 'note': 'Формируется обязательно'},
                {'class': 'Страхование ответственности', 'loss_ratio': Decimal('27.0'), 'threshold': Decimal('50'),
                 'deviation': Decimal('0'), 'reserve': Decimal('50000000'), 'note': 'Накопленный резерв'},
                {'class': 'Прочие классы', 'loss_ratio': Decimal('48.5'), 'threshold': Decimal('55'),
                 'deviation': Decimal('0'), 'reserve': Decimal('50000000'), 'note': 'Накопленный резерв'},
            ],
            'total': Decimal('250000000'),
            'formula': 'СР = max(0, (LR_факт - LR_норм) × Премия × 0.1)',
            'note': 'Согласно п.25 Постановления АРРФР №304'
        }

    def _export_to_excel(self, report_code: str, metadata: ReportMetadata, data: Dict) -> tuple:
        """Полноценный экспорт в Excel с форматированием"""
        if not OPENPYXL_AVAILABLE:
            return self._export_to_json(report_code, metadata, data)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = report_code

        # Стили
        header_font = Font(name='Arial', size=14, bold=True)
        title_font = Font(name='Arial', size=11, bold=True)
        normal_font = Font(name='Arial', size=10)
        small_font = Font(name='Arial', size=9)

        header_fill = PatternFill(start_color='1e3a5f', end_color='1e3a5f', fill_type='solid')
        subheader_fill = PatternFill(start_color='2c5282', end_color='2c5282', fill_type='solid')
        total_fill = PatternFill(start_color='e2e8f0', end_color='e2e8f0', fill_type='solid')
        success_fill = PatternFill(start_color='c6efce', end_color='c6efce', fill_type='solid')
        warning_fill = PatternFill(start_color='ffeb9c', end_color='ffeb9c', fill_type='solid')

        white_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')

        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        number_format = '#,##0'
        decimal_format = '#,##0.00'
        percent_format = '0.00%'

        # === ЗАГОЛОВОК ДОКУМЕНТА ===
        row = 1

        # Шапка с постановлением
        ws.merge_cells(f'A{row}:H{row}')
        ws[f'A{row}'] = metadata.regulation
        ws[f'A{row}'].font = small_font
        ws[f'A{row}'].alignment = Alignment(horizontal='right')
        row += 1

        # Название формы
        ws.merge_cells(f'A{row}:H{row}')
        ws[f'A{row}'] = f'Форма {report_code}'
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        row += 1

        # Название отчёта
        ws.merge_cells(f'A{row}:H{row}')
        ws[f'A{row}'] = metadata.report_name
        ws[f'A{row}'].font = title_font
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        row += 2

        # Информация о компании
        ws[f'A{row}'] = 'Наименование страховой организации:'
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws.merge_cells(f'B{row}:D{row}')
        ws[f'B{row}'] = metadata.company_name
        ws[f'F{row}'] = 'БИН:'
        ws[f'F{row}'].font = Font(bold=True, size=10)
        ws[f'G{row}'] = metadata.company_bin
        row += 1

        ws[f'A{row}'] = 'Отчётный период:'
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'B{row}'] = f'{metadata.period_start.strftime("%d.%m.%Y")} - {metadata.period_end.strftime("%d.%m.%Y")}'
        ws[f'F{row}'] = 'Дата формирования:'
        ws[f'F{row}'].font = Font(bold=True, size=10)
        ws[f'G{row}'] = metadata.generated_at.strftime('%d.%m.%Y %H:%M')
        row += 2

        # === ОСНОВНОЕ СОДЕРЖИМОЕ В ЗАВИСИМОСТИ ОТ ФОРМЫ ===

        if report_code == '1-SK':
            row = self._write_1sk_excel(ws, data, row, header_fill, white_font, total_fill, border, number_format)
        elif report_code == '2-SK':
            row = self._write_2sk_excel(ws, data, row, header_fill, white_font, total_fill, border, number_format)
        elif report_code == '3-SK':
            row = self._write_3sk_excel(ws, data, row, header_fill, white_font, total_fill, border, number_format, success_fill)
        elif report_code == '4-SK':
            row = self._write_4sk_excel(ws, data, row, header_fill, white_font, total_fill, border, number_format)
        elif report_code == 'P86-1':
            row = self._write_p861_excel(ws, data, row, header_fill, subheader_fill, white_font, total_fill, success_fill, border, number_format)
        elif report_code == 'P86-2':
            row = self._write_p862_excel(ws, data, row, header_fill, white_font, total_fill, success_fill, border, number_format)
        elif report_code == 'P86-3':
            row = self._write_p863_excel(ws, data, row, header_fill, white_font, total_fill, success_fill, border, number_format)
        elif report_code == 'R304-1':
            row = self._write_r3041_excel(ws, data, row, header_fill, white_font, total_fill, border, number_format)
        elif report_code == 'R304-2':
            row = self._write_r3042_excel(ws, data, row, header_fill, white_font, total_fill, border, number_format)
        elif report_code == 'R304-3':
            row = self._write_r3043_excel(ws, data, row, header_fill, subheader_fill, white_font, total_fill, border, number_format)
        elif report_code == 'R304-4':
            row = self._write_r3044_excel(ws, data, row, header_fill, white_font, total_fill, border, number_format)

        # === ПОДПИСИ ===
        row += 2
        ws[f'A{row}'] = f'Руководитель: _________________________ / {self.company_info["director"]}'
        row += 1
        ws[f'A{row}'] = f'Главный бухгалтер: _________________________ / {self.company_info["accountant"]}'
        row += 1
        ws[f'A{row}'] = 'М.П.'
        row += 2
        ws[f'A{row}'] = f'Дата составления: {metadata.generated_at.strftime("%d.%m.%Y")}'

        # Настройка ширины колонок
        column_widths = {'A': 8, 'B': 45, 'C': 18, 'D': 18, 'E': 18, 'F': 15, 'G': 15, 'H': 15}
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Сохранение
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f'{report_code}_{metadata.period}.xlsx'
        return output.getvalue(), filename

    def _write_1sk_excel(self, ws, data, row, header_fill, white_font, total_fill, border, number_format):
        """Запись формы 1-СК в Excel"""
        # Заголовок таблицы
        headers = ['№ п/п', 'Показатель', 'Отчётный период (тыс.₸)', 'Аналогичный период прошлого года (тыс.₸)', 'Изменение (%)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = white_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        row += 1

        # Данные
        for item in data.get('sections', []):
            ws.cell(row=row, column=1, value=item.get('row', '')).border = border
            ws.cell(row=row, column=2, value=item.get('name', '')).border = border

            value = item.get('value', 0)
            prev_value = item.get('prev_value', 0)

            # Форматирование числовых значений
            if isinstance(value, Decimal):
                if value < 100:  # Процент или коэффициент
                    cell = ws.cell(row=row, column=3, value=float(value))
                    cell.number_format = '0.00'
                else:
                    cell = ws.cell(row=row, column=3, value=int(value / 1000))
                    cell.number_format = number_format
            else:
                cell = ws.cell(row=row, column=3, value=value)
                if isinstance(value, int) and value > 100:
                    cell.number_format = number_format
            cell.border = border

            if isinstance(prev_value, Decimal):
                if prev_value < 100:
                    cell = ws.cell(row=row, column=4, value=float(prev_value))
                    cell.number_format = '0.00'
                else:
                    cell = ws.cell(row=row, column=4, value=int(prev_value / 1000))
                    cell.number_format = number_format
            else:
                cell = ws.cell(row=row, column=4, value=prev_value)
                if isinstance(prev_value, int) and prev_value > 100:
                    cell.number_format = number_format
            cell.border = border

            # Изменение в %
            if prev_value and prev_value != 0:
                if isinstance(value, Decimal) and isinstance(prev_value, Decimal):
                    change = float((value - prev_value) / prev_value * 100)
                elif isinstance(value, (int, float)) and isinstance(prev_value, (int, float)):
                    change = (value - prev_value) / prev_value * 100
                else:
                    change = 0
                cell = ws.cell(row=row, column=5, value=change)
                cell.number_format = '+0.0%;-0.0%'
            else:
                ws.cell(row=row, column=5, value='-')
            ws.cell(row=row, column=5).border = border

            # Выделение итоговых строк
            if item.get('row', '').startswith(('1', '2')) and '.' not in str(item.get('row', '')):
                for col in range(1, 6):
                    ws.cell(row=row, column=col).fill = total_fill
                    ws.cell(row=row, column=col).font = Font(bold=True)

            row += 1

        return row

    def _write_2sk_excel(self, ws, data, row, header_fill, white_font, total_fill, border, number_format):
        """Запись формы 2-СК в Excel"""
        headers = ['№', 'Класс страхования', 'Премии брутто (тыс.₸)', 'Премии нетто (тыс.₸)',
                   'Выплаты брутто (тыс.₸)', 'Выплаты нетто (тыс.₸)', 'Кол-во договоров', 'Убыточность (%)']

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = white_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        row += 1

        for item in data.get('sections', []):
            is_total = item.get('is_total', False) or item.get('class_name') == 'ИТОГО'

            ws.cell(row=row, column=1, value=item.get('row', '')).border = border
            ws.cell(row=row, column=2, value=item.get('class_name', '')).border = border

            cell = ws.cell(row=row, column=3, value=int(item.get('premiums_gross', 0) / 1000))
            cell.number_format = number_format
            cell.border = border

            cell = ws.cell(row=row, column=4, value=int(item.get('premiums_net', 0) / 1000))
            cell.number_format = number_format
            cell.border = border

            cell = ws.cell(row=row, column=5, value=int(item.get('claims_gross', 0) / 1000))
            cell.number_format = number_format
            cell.border = border

            cell = ws.cell(row=row, column=6, value=int(item.get('claims_net', 0) / 1000))
            cell.number_format = number_format
            cell.border = border

            cell = ws.cell(row=row, column=7, value=item.get('contracts', 0))
            cell.number_format = number_format
            cell.border = border

            cell = ws.cell(row=row, column=8, value=float(item.get('loss_ratio', 0)))
            cell.number_format = '0.00'
            cell.border = border

            if is_total:
                for col in range(1, 9):
                    ws.cell(row=row, column=col).fill = total_fill
                    ws.cell(row=row, column=col).font = Font(bold=True)

            row += 1

        return row

    def _write_3sk_excel(self, ws, data, row, header_fill, white_font, total_fill, border, number_format, success_fill):
        """Запись формы 3-СК в Excel"""
        headers = ['№', 'Вид актива', 'Сумма (тыс.₸)', 'Доля (%)', 'Норматив', 'Статус', 'Примечание']

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = white_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        row += 1

        for item in data.get('sections', []):
            is_total = 'ИТОГО' in item.get('asset_type', '')

            ws.cell(row=row, column=1, value=item.get('row', '')).border = border
            ws.cell(row=row, column=2, value=item.get('asset_type', '')).border = border

            cell = ws.cell(row=row, column=3, value=int(item.get('amount', 0) / 1000))
            cell.number_format = number_format
            cell.border = border

            cell = ws.cell(row=row, column=4, value=float(item.get('share', 0)))
            cell.number_format = '0.00'
            cell.border = border

            limit = item.get('limit')
            if limit:
                ws.cell(row=row, column=5, value=f'≤{float(limit)}%').border = border
            else:
                ws.cell(row=row, column=5, value='').border = border

            status = item.get('status', '')
            cell = ws.cell(row=row, column=6, value=status)
            cell.border = border
            if status == 'OK':
                cell.fill = success_fill

            ws.cell(row=row, column=7, value=item.get('note', '')).border = border

            if is_total:
                for col in range(1, 8):
                    ws.cell(row=row, column=col).fill = total_fill
                    ws.cell(row=row, column=col).font = Font(bold=True)

            row += 1

        return row

    def _write_4sk_excel(self, ws, data, row, header_fill, white_font, total_fill, border, number_format):
        """Запись формы 4-СК в Excel"""
        # Раздел 1 - Сводные данные
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'] = 'Раздел 1. Сводные данные по перестрахованию'
        ws[f'A{row}'].font = Font(bold=True, color='0066CC')
        row += 1

        headers = ['№', 'Показатель', 'Сумма (тыс.₸)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = white_font
            cell.border = border
        row += 1

        for item in data['sections'].get('summary', []):
            ws.cell(row=row, column=1, value=item.get('row', '')).border = border
            ws.cell(row=row, column=2, value=item.get('name', '')).border = border
            value = item.get('value', 0)
            if isinstance(value, Decimal) and value < 1:
                cell = ws.cell(row=row, column=3, value=float(value))
                cell.number_format = '0.00'
            else:
                cell = ws.cell(row=row, column=3, value=int(value / 1000) if isinstance(value, Decimal) else value)
                cell.number_format = number_format
            cell.border = border
            row += 1

        row += 1

        # Раздел 2 - По классам
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'] = 'Раздел 2. Перестрахование по классам страхования'
        ws[f'A{row}'].font = Font(bold=True, color='0066CC')
        row += 1

        headers = ['Класс страхования', 'Переданные премии (тыс.₸)', 'Доля перестраховщика в выплатах (тыс.₸)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = white_font
            cell.border = border
        row += 1

        for item in data['sections'].get('by_class', []):
            ws.cell(row=row, column=1, value=item.get('class', '')).border = border
            cell = ws.cell(row=row, column=2, value=int(item.get('ceded', 0) / 1000))
            cell.number_format = number_format
            cell.border = border
            cell = ws.cell(row=row, column=3, value=int(item.get('recovered', 0) / 1000))
            cell.number_format = number_format
            cell.border = border
            row += 1

        row += 1

        # Раздел 3 - Перестраховщики
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'] = 'Раздел 3. Информация о перестраховщиках'
        ws[f'A{row}'].font = Font(bold=True, color='0066CC')
        row += 1

        headers = ['Наименование', 'Страна', 'Рейтинг', 'Доля (%)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = white_font
            cell.border = border
        row += 1

        for item in data['sections'].get('reinsurers', []):
            ws.cell(row=row, column=1, value=item.get('name', '')).border = border
            ws.cell(row=row, column=2, value=item.get('country', '')).border = border
            ws.cell(row=row, column=3, value=item.get('rating', '')).border = border
            ws.cell(row=row, column=4, value=float(item.get('share', 0))).border = border
            row += 1

        return row

    def _write_p861_excel(self, ws, data, row, header_fill, subheader_fill, white_font, total_fill, success_fill, border, number_format):
        """Запись формы П-86/1 в Excel"""
        sections = data.get('sections', {})

        # Раздел 1 - ФМП
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'] = 'Раздел 1. Расчёт фактической маржи платёжеспособности (ФМП)'
        ws[f'A{row}'].font = Font(bold=True, color='0066CC', size=11)
        row += 1

        headers = ['№ п/п', 'Показатель', 'Сумма (тыс.₸)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = white_font
            cell.border = border
        row += 1

        for item in sections.get('fmp_section', []):
            ws.cell(row=row, column=1, value=item.get('row', '')).border = border
            ws.cell(row=row, column=2, value=item.get('name', '')).border = border
            cell = ws.cell(row=row, column=3, value=int(item.get('value', 0) / 1000))
            cell.number_format = number_format
            cell.border = border

            if item.get('is_total'):
                for col in range(1, 4):
                    ws.cell(row=row, column=col).fill = total_fill
                    ws.cell(row=row, column=col).font = Font(bold=True)
            row += 1

        row += 1

        # Раздел 2 - НМП
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'] = 'Раздел 2. Расчёт нормативной маржи платёжеспособности (НМП)'
        ws[f'A{row}'].font = Font(bold=True, color='0066CC', size=11)
        row += 1

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = white_font
            cell.border = border
        row += 1

        for item in sections.get('nmp_section', []):
            ws.cell(row=row, column=1, value=item.get('row', '')).border = border
            ws.cell(row=row, column=2, value=item.get('name', '')).border = border
            value = item.get('value', 0)
            if isinstance(value, Decimal) and value < 1:
                cell = ws.cell(row=row, column=3, value=float(value))
                cell.number_format = '0.00'
            else:
                cell = ws.cell(row=row, column=3, value=int(value / 1000))
                cell.number_format = number_format
            cell.border = border

            if item.get('is_total'):
                for col in range(1, 4):
                    ws.cell(row=row, column=col).fill = total_fill
                    ws.cell(row=row, column=col).font = Font(bold=True)
            row += 1

        row += 1

        # Раздел 3 - МГФ
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'] = 'Раздел 3. Расчёт минимального гарантийного фонда (МГФ)'
        ws[f'A{row}'].font = Font(bold=True, color='0066CC', size=11)
        row += 1

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = white_font
            cell.border = border
        row += 1

        for item in sections.get('mgf_section', []):
            ws.cell(row=row, column=1, value=item.get('row', '')).border = border
            ws.cell(row=row, column=2, value=item.get('name', '')).border = border
            cell = ws.cell(row=row, column=3, value=int(item.get('value', 0) / 1000))
            cell.number_format = number_format
            cell.border = border

            if item.get('is_total'):
                for col in range(1, 4):
                    ws.cell(row=row, column=col).fill = total_fill
                    ws.cell(row=row, column=col).font = Font(bold=True)
            row += 1

        row += 1

        # Итоговый расчёт
        result = sections.get('result', {})
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'] = 'ИТОГОВЫЙ РАСЧЁТ'
        ws[f'A{row}'].font = Font(bold=True, color='0066CC', size=11)
        row += 1

        ws.cell(row=row, column=1, value='').border = border
        ws.cell(row=row, column=2, value='Коэффициент достаточности (ФМП/НМП)').border = border
        cell = ws.cell(row=row, column=3, value=f'{float(result.get("ratio", 0)) * 100:.1f}%')
        cell.border = border
        cell.font = Font(bold=True, size=12)
        cell.fill = success_fill if float(result.get('ratio', 0)) >= 1 else PatternFill(start_color='ffc7ce', end_color='ffc7ce', fill_type='solid')
        row += 1

        ws.cell(row=row, column=1, value='').border = border
        ws.cell(row=row, column=2, value='Превышение ФМП над НМП').border = border
        cell = ws.cell(row=row, column=3, value=int(result.get('surplus', 0) / 1000))
        cell.number_format = number_format
        cell.border = border
        row += 1

        ws.cell(row=row, column=1, value='').border = border
        ws.cell(row=row, column=2, value='СТАТУС').border = border
        ws.cell(row=row, column=2).font = Font(bold=True)
        cell = ws.cell(row=row, column=3, value=result.get('status', ''))
        cell.border = border
        cell.font = Font(bold=True)
        cell.fill = success_fill if result.get('status') == 'СООТВЕТСТВУЕТ' else PatternFill(start_color='ffc7ce', end_color='ffc7ce', fill_type='solid')
        row += 1

        return row

    def _write_p862_excel(self, ws, data, row, header_fill, white_font, total_fill, success_fill, border, number_format):
        """Запись формы П-86/2 в Excel"""
        headers = ['№ п/п', 'Показатель', 'Сумма (тыс.₸)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = white_font
            cell.border = border
        row += 1

        for item in data['sections'].get('calculation', []):
            ws.cell(row=row, column=1, value=item.get('row', '')).border = border
            ws.cell(row=row, column=2, value=item.get('name', '')).border = border
            cell = ws.cell(row=row, column=3, value=int(item.get('value', 0) / 1000))
            cell.number_format = number_format
            cell.border = border

            if item.get('is_total'):
                for col in range(1, 4):
                    ws.cell(row=row, column=col).fill = total_fill
                    ws.cell(row=row, column=col).font = Font(bold=True)
            row += 1

        row += 1
        ws.cell(row=row, column=2, value='СТАТУС').font = Font(bold=True)
        cell = ws.cell(row=row, column=3, value=data['sections'].get('status', ''))
        cell.font = Font(bold=True)
        cell.fill = success_fill if data['sections'].get('status') == 'СООТВЕТСТВУЕТ' else PatternFill(start_color='ffc7ce', end_color='ffc7ce', fill_type='solid')

        return row + 1

    def _write_p863_excel(self, ws, data, row, header_fill, white_font, total_fill, success_fill, border, number_format):
        """Запись формы П-86/3 в Excel"""
        headers = ['Код', 'Пруденциальный норматив', 'Нормативное значение', 'Фактическое значение', 'Статус']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = white_font
            cell.border = border
        row += 1

        for item in data['sections'].get('norms', []):
            ws.cell(row=row, column=1, value=item.get('code', '')).border = border
            ws.cell(row=row, column=2, value=item.get('name', '')).border = border
            ws.cell(row=row, column=3, value=item.get('norm', '')).border = border
            cell = ws.cell(row=row, column=4, value=float(item.get('actual', 0)))
            cell.number_format = '0.00'
            cell.border = border

            status = item.get('status', '')
            cell = ws.cell(row=row, column=5, value=status)
            cell.border = border
            cell.fill = success_fill if status == 'OK' else PatternFill(start_color='ffc7ce', end_color='ffc7ce', fill_type='solid')
            row += 1

        return row

    def _write_r3041_excel(self, ws, data, row, header_fill, white_font, total_fill, border, number_format):
        """Запись формы Р-304/1 в Excel"""
        # Описание метода
        ws.merge_cells(f'A{row}:G{row}')
        ws[f'A{row}'] = 'Метод расчёта: Pro-rata temporis (1/365) согласно п.12 Постановления АРРФР №304'
        ws[f'A{row}'].font = Font(italic=True, color='666666')
        row += 2

        headers = ['№', 'Класс страхования', 'Премии (тыс.₸)', 'Начало', 'Окончание', 'Доля незараб. (%)', 'РНПП (тыс.₸)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = white_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        row += 1

        for item in data.get('sections', []):
            is_total = item.get('is_total', False)

            ws.cell(row=row, column=1, value=item.get('row', '')).border = border
            ws.cell(row=row, column=2, value=item.get('class', '')).border = border

            cell = ws.cell(row=row, column=3, value=int(item.get('premiums', 0) / 1000))
            cell.number_format = number_format
            cell.border = border

            ws.cell(row=row, column=4, value=item.get('inception_date', '')).border = border
            ws.cell(row=row, column=5, value=item.get('expiry_date', '')).border = border

            cell = ws.cell(row=row, column=6, value=float(item.get('unearned_ratio', 0)))
            cell.number_format = '0.00'
            cell.border = border

            cell = ws.cell(row=row, column=7, value=int(item.get('rnpp', 0) / 1000))
            cell.number_format = number_format
            cell.border = border

            if is_total:
                for col in range(1, 8):
                    ws.cell(row=row, column=col).fill = total_fill
                    ws.cell(row=row, column=col).font = Font(bold=True)

            row += 1

        # Формула
        row += 1
        ws.merge_cells(f'A{row}:G{row}')
        ws[f'A{row}'] = 'Формула: РНПП = Премия × (Количество дней до окончания / Общий срок договора в днях)'
        ws[f'A{row}'].font = Font(italic=True)

        return row + 1

    def _write_r3042_excel(self, ws, data, row, header_fill, white_font, total_fill, border, number_format):
        """Запись формы Р-304/2 в Excel"""
        headers = ['№', 'Класс страхования', 'Кол-во убытков', 'Оценка брутто (тыс.₸)', 'Оценка нетто (тыс.₸)', 'РРУНУ (тыс.₸)', 'РЗУ итого (тыс.₸)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = white_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        row += 1

        for item in data.get('sections', []):
            is_total = item.get('is_total', False)

            ws.cell(row=row, column=1, value=item.get('row', '')).border = border
            ws.cell(row=row, column=2, value=item.get('class', '')).border = border

            cell = ws.cell(row=row, column=3, value=item.get('claims_count', 0))
            cell.number_format = number_format
            cell.border = border

            cell = ws.cell(row=row, column=4, value=int(item.get('estimated_gross', 0) / 1000))
            cell.number_format = number_format
            cell.border = border

            cell = ws.cell(row=row, column=5, value=int(item.get('estimated_net', 0) / 1000))
            cell.number_format = number_format
            cell.border = border

            cell = ws.cell(row=row, column=6, value=int(item.get('rpzu', 0) / 1000))
            cell.number_format = number_format
            cell.border = border

            cell = ws.cell(row=row, column=7, value=int(item.get('total', 0) / 1000))
            cell.number_format = number_format
            cell.border = border

            if is_total:
                for col in range(1, 8):
                    ws.cell(row=row, column=col).fill = total_fill
                    ws.cell(row=row, column=col).font = Font(bold=True)

            row += 1

        # Примечание
        row += 1
        ws.merge_cells(f'A{row}:G{row}')
        ws[f'A{row}'] = 'РРУНУ (Резерв расходов на урегулирование) = 10% от оценки убытков согласно п.18 Постановления №304'
        ws[f'A{row}'].font = Font(italic=True)

        return row + 1

    def _write_r3043_excel(self, ws, data, row, header_fill, subheader_fill, white_font, total_fill, border, number_format):
        """Запись формы Р-304/3 в Excel - IBNR Chain-Ladder"""
        sections = data.get('sections', {})

        # Треугольник развития
        ws.merge_cells(f'A{row}:H{row}')
        ws[f'A{row}'] = 'Треугольник развития убытков (тыс.тенге)'
        ws[f'A{row}'].font = Font(bold=True, color='0066CC', size=11)
        row += 1

        headers = ['Год', 'Dev 0', 'Dev 1', 'Dev 2', 'Dev 3', 'Dev 4', 'Ultimate', 'IBNR']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = white_font
            cell.border = border
        row += 1

        for item in sections.get('triangle', []):
            ws.cell(row=row, column=1, value=item.get('year', '')).border = border

            for col, key in enumerate(['dev0', 'dev1', 'dev2', 'dev3', 'dev4'], 2):
                value = item.get(key)
                cell = ws.cell(row=row, column=col)
                if value is not None:
                    cell.value = value
                    cell.number_format = number_format
                else:
                    cell.value = '-'
                    cell.fill = PatternFill(start_color='fff3cd', end_color='fff3cd', fill_type='solid')
                cell.border = border

            cell = ws.cell(row=row, column=7, value=item.get('ultimate', 0))
            cell.number_format = number_format
            cell.border = border

            cell = ws.cell(row=row, column=8, value=item.get('ibnr', 0))
            cell.number_format = number_format
            cell.border = border
            if item.get('ibnr', 0) > 0:
                cell.fill = PatternFill(start_color='e7f3ff', end_color='e7f3ff', fill_type='solid')

            row += 1

        row += 1

        # LDF
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = 'Коэффициенты развития (LDF - Loss Development Factors)'
        ws[f'A{row}'].font = Font(bold=True, color='0066CC', size=11)
        row += 1

        headers = ['Период', '0→1', '1→2', '2→3', '3→4', '4→Ult']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = subheader_fill
            cell.font = white_font
            cell.border = border
        row += 1

        ldf = sections.get('ldf', {})
        ws.cell(row=row, column=1, value='LDF').border = border
        ws.cell(row=row, column=2, value=float(ldf.get('0_1', 0))).border = border
        ws.cell(row=row, column=3, value=float(ldf.get('1_2', 0))).border = border
        ws.cell(row=row, column=4, value=float(ldf.get('2_3', 0))).border = border
        ws.cell(row=row, column=5, value=float(ldf.get('3_4', 0))).border = border
        ws.cell(row=row, column=6, value=float(ldf.get('4_ult', 0))).border = border
        row += 1

        cdf = sections.get('cdf', {})
        ws.cell(row=row, column=1, value='CDF (накопл.)').border = border
        ws.cell(row=row, column=2, value=float(cdf.get('0_1', 0))).border = border
        ws.cell(row=row, column=3, value=float(cdf.get('1_2', 0))).border = border
        ws.cell(row=row, column=4, value=float(cdf.get('2_3', 0))).border = border
        ws.cell(row=row, column=5, value=float(cdf.get('3_4', 0))).border = border
        ws.cell(row=row, column=6, value=float(cdf.get('4_ult', 0))).border = border
        row += 2

        # Итого IBNR
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'] = 'Итого IBNR по классам страхования'
        ws[f'A{row}'].font = Font(bold=True, color='0066CC', size=11)
        row += 1

        total_ibnr = sections.get('total_ibnr', {})
        for class_name, key in [('ОГПО ВТС', 'ogpo_vts'), ('КАСКО', 'kasko'), ('Имущество', 'property'), ('Прочие', 'other')]:
            ws.cell(row=row, column=1, value=class_name).border = border
            cell = ws.cell(row=row, column=2, value=int(total_ibnr.get(key, 0) / 1000))
            cell.number_format = number_format
            cell.border = border
            row += 1

        ws.cell(row=row, column=1, value='ИТОГО IBNR').border = border
        ws.cell(row=row, column=1).font = Font(bold=True)
        cell = ws.cell(row=row, column=2, value=int(total_ibnr.get('total', 0) / 1000))
        cell.number_format = number_format
        cell.border = border
        cell.font = Font(bold=True)
        cell.fill = total_fill

        return row + 1

    def _write_r3044_excel(self, ws, data, row, header_fill, white_font, total_fill, border, number_format):
        """Запись формы Р-304/4 в Excel"""
        sections = data.get('sections', {})

        headers = ['Класс страхования', 'Убыточность факт (%)', 'Порог (%)', 'Превышение', 'Резерв (тыс.₸)', 'Примечание']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = white_font
            cell.border = border
        row += 1

        for item in sections.get('by_class', []):
            ws.cell(row=row, column=1, value=item.get('class', '')).border = border

            cell = ws.cell(row=row, column=2, value=float(item.get('loss_ratio', 0)))
            cell.number_format = '0.0'
            cell.border = border

            cell = ws.cell(row=row, column=3, value=float(item.get('threshold', 0)))
            cell.number_format = '0.0'
            cell.border = border

            cell = ws.cell(row=row, column=4, value=float(item.get('deviation', 0)))
            cell.number_format = '0.0'
            cell.border = border

            cell = ws.cell(row=row, column=5, value=int(item.get('reserve', 0) / 1000))
            cell.number_format = number_format
            cell.border = border

            ws.cell(row=row, column=6, value=item.get('note', '')).border = border
            row += 1

        # Итого
        ws.cell(row=row, column=1, value='ИТОГО').border = border
        ws.cell(row=row, column=1).font = Font(bold=True)
        for col in range(2, 5):
            ws.cell(row=row, column=col, value='').border = border
        cell = ws.cell(row=row, column=5, value=int(sections.get('total', 0) / 1000))
        cell.number_format = number_format
        cell.border = border
        cell.font = Font(bold=True)
        cell.fill = total_fill
        ws.cell(row=row, column=6, value='').border = border
        row += 2

        # Формула
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = f'Формула: {sections.get("formula", "")}'
        ws[f'A{row}'].font = Font(italic=True)
        row += 1
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = sections.get('note', '')
        ws[f'A{row}'].font = Font(italic=True, color='666666')

        return row + 1

    def _export_to_pdf(self, report_code: str, metadata: ReportMetadata, data: Dict) -> tuple:
        """Экспорт в PDF"""
        if not REPORTLAB_AVAILABLE:
            return self._export_to_json(report_code, metadata, data)

        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4)
        elements = []

        styles = getSampleStyleSheet()

        # Title
        elements.append(Paragraph(f'<b>Форма {report_code}</b>', styles['Title']))
        elements.append(Paragraph(f'<b>{metadata.report_name}</b>', styles['Heading2']))
        elements.append(Paragraph(metadata.regulation, styles['Normal']))
        elements.append(Spacer(1, 20))

        # Company info
        elements.append(Paragraph(f'Наименование: {metadata.company_name}', styles['Normal']))
        elements.append(Paragraph(f'БИН: {metadata.company_bin}', styles['Normal']))
        elements.append(Paragraph(f'Период: {metadata.period_start.strftime("%d.%m.%Y")} - {metadata.period_end.strftime("%d.%m.%Y")}', styles['Normal']))
        elements.append(Spacer(1, 20))

        # Simplified data table
        if 'sections' in data:
            sections = data['sections']
            if isinstance(sections, list) and len(sections) > 0:
                # Create table from list data
                if 'name' in sections[0]:
                    table_data = [['№', 'Показатель', 'Значение']]
                    for item in sections[:15]:  # Limit rows
                        table_data.append([
                            str(item.get('row', '')),
                            str(item.get('name', item.get('class_name', item.get('class', ''))))[:40],
                            str(item.get('value', item.get('rnpp', item.get('total', ''))))
                        ])
                elif 'class_name' in sections[0]:
                    table_data = [['Класс', 'Премии', 'Выплаты']]
                    for item in sections[:15]:
                        table_data.append([
                            str(item.get('class_name', ''))[:30],
                            str(int(item.get('premiums_gross', 0) / 1000)),
                            str(int(item.get('claims_gross', 0) / 1000))
                        ])
                else:
                    table_data = [['Класс', 'Премии', 'РНПП']]
                    for item in sections[:15]:
                        table_data.append([
                            str(item.get('class', ''))[:30],
                            str(int(item.get('premiums', 0) / 1000)),
                            str(int(item.get('rnpp', item.get('total', 0)) / 1000))
                        ])

                table = Table(table_data, colWidths=[50, 250, 100])
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a5f')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                elements.append(table)

        elements.append(Spacer(1, 30))
        elements.append(Paragraph(f'Дата формирования: {metadata.generated_at.strftime("%d.%m.%Y %H:%M")}', styles['Normal']))

        doc.build(elements)
        output.seek(0)

        filename = f'{report_code}_{metadata.period}.pdf'
        return output.getvalue(), filename

    def _export_to_json(self, report_code: str, metadata: ReportMetadata, data: Dict) -> tuple:
        """Экспорт в JSON"""
        result = {
            'metadata': {
                'report_code': metadata.report_code,
                'report_name': metadata.report_name,
                'report_name_kz': metadata.report_name_kz,
                'regulation': metadata.regulation,
                'period': metadata.period,
                'period_start': metadata.period_start.isoformat(),
                'period_end': metadata.period_end.isoformat(),
                'company_name': metadata.company_name,
                'company_bin': metadata.company_bin,
                'generated_at': metadata.generated_at.isoformat()
            },
            'data': self._serialize_data(data)
        }

        content = json.dumps(result, ensure_ascii=False, indent=2).encode('utf-8')
        filename = f'{report_code}_{metadata.period}.json'
        return content, filename

    def _serialize_data(self, data: Any) -> Any:
        """Сериализация данных для JSON"""
        if isinstance(data, Decimal):
            return str(data)
        elif isinstance(data, date):
            return data.isoformat()
        elif isinstance(data, dict):
            return {k: self._serialize_data(v) for k, v in data.items()}
        elif isinstance(data, list):
            return [self._serialize_data(item) for item in data]
        else:
            return data


# Singleton instance
report_generator_service = ReportGeneratorService()
